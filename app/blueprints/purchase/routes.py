"""采购管理路由"""
from flask import render_template, request, flash, redirect, url_for, jsonify, Response
from flask_login import login_required, current_user
from app.extensions import db
from app.blueprints.purchase import purchase_bp
from app.blueprints.purchase.forms import PurchaseOrderForm
from app.models.purchase import PurchaseOrder, PurchaseOrderItem, PurchasePriceHistory
from app.models.biz import Partner, Product
from app.models.stock import Warehouse
from app.services.purchase_service import PurchaseService
from app.utils.decorators import permission_required
import csv
import io
from datetime import datetime


@purchase_bp.route('/')
@login_required
def index():
    """采购订单列表"""
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    supplier_id = request.args.get('supplier_id', 0, type=int)
    export_format = request.args.get('export', '')
    
    query = PurchaseOrder.query
    
    if status:
        query = query.filter_by(status=status)
    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)
    
    # 导出功能
    if export_format in ['excel', 'csv']:
        orders_to_export = query.order_by(PurchaseOrder.created_at.desc()).all()
        return export_purchase_orders(orders_to_export, export_format)
    
    pagination = query.order_by(PurchaseOrder.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False
    )
    
    suppliers = Partner.query.filter(
        Partner.type.in_(['supplier', 'both']),
        Partner.is_deleted == False
    ).all()
    
    # 计算各状态数量（用于流程图显示）
    status_counts = {
        'draft': PurchaseOrder.query.filter_by(status='draft').count(),
        'pending': PurchaseOrder.query.filter_by(status='pending').count(),
        'approved': PurchaseOrder.query.filter_by(status='approved').count(),
        'received': PurchaseOrder.query.filter_by(status='received').count(),
        'cancelled': PurchaseOrder.query.filter_by(status='cancelled').count(),
    }
    
    return render_template('purchase/index.html',
                         orders=pagination.items,
                         pagination=pagination,
                         suppliers=suppliers,
                         current_status=status,
                         current_supplier=supplier_id,
                         status_counts=status_counts)


def export_purchase_orders(orders, format_type):
    """导出采购订单"""
    # 准备数据
    data = []
    for order in orders:
        status_map = {
            'draft': '草稿',
            'pending': '待审批',
            'approved': '已审批',
            'received': '已入库',
            'cancelled': '已取消'
        }
        data.append({
            '采购单号': order.po_no,
            '供应商': order.supplier.name if order.supplier else '',
            '入库仓库': order.warehouse.name if order.warehouse else '',
            '商品数量': len(order.items),
            '采购总额': f'{order.total_amount:.2f}',
            '状态': status_map.get(order.status, order.status),
            '创建时间': order.created_at.strftime('%Y-%m-%d %H:%M'),
            '备注': order.remark or ''
        })
    
    if format_type == 'csv':
        # 导出CSV
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        response = Response(
            '\ufeff' + output.getvalue(),  # 添加BOM以支持Excel打开
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename=purchase_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            }
        )
        return response
    
    else:
        # 导出Excel (简化版，使用CSV格式但扩展名为xlsx)
        # 如果需要真正的Excel格式，需要安装openpyxl
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "采购订单"
            
            # 写入表头
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 写入数据
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, key in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename=purchase_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                }
            )
            return response
        except ImportError:
            # 如果没有openpyxl，则导出CSV
            output = io.StringIO()
            if data:
                writer = csv.DictWriter(output, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
            
            response = Response(
                '\ufeff' + output.getvalue(),
                mimetype='text/csv; charset=utf-8',
                headers={
                    'Content-Disposition': f'attachment; filename=purchase_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                }
            )
            return response


@purchase_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """创建采购订单"""
    form = PurchaseOrderForm()
    
    # 加载下拉选项
    suppliers = Partner.query.filter(
        Partner.type.in_(['supplier', 'both']),
        Partner.is_deleted == False
    ).all()
    form.supplier_id.choices = [(s.id, s.name) for s in suppliers]
    
    warehouses = Warehouse.query.filter_by(is_deleted=False).all()
    form.warehouse_id.choices = [(w.id, w.name) for w in warehouses]
    
    if request.method == 'POST':
        # 获取商品数据
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        unit_prices = request.form.getlist('unit_price[]')
        
        if not product_ids:
            flash('请添加采购商品', 'danger')
            return render_template('purchase/create.html', form=form)
        
        items = []
        for i, pid in enumerate(product_ids):
            if pid and quantities[i]:
                items.append({
                    'product_id': int(pid),
                    'quantity': int(quantities[i]),
                    'unit_price': float(unit_prices[i]) if unit_prices[i] else 0
                })
        
        if not items:
            flash('请添加有效的采购商品', 'danger')
            return render_template('purchase/create.html', form=form)
        
        success, result = PurchaseService.create_purchase_order(
            supplier_id=form.supplier_id.data,
            warehouse_id=form.warehouse_id.data,
            items=items,
            user=current_user,
            remark=form.remark.data
        )
        
        if success:
            flash(f'采购订单 {result.po_no} 创建成功', 'success')
            return redirect(url_for('purchase.detail', po_id=result.id))
        else:
            flash(result, 'danger')
    
    # 加载商品列表
    products = Product.query.filter_by(is_deleted=False).all()
    
    return render_template('purchase/create.html', form=form, products=products)


@purchase_bp.route('/<int:po_id>')
@login_required
def detail(po_id):
    """采购订单详情"""
    po = PurchaseOrder.query.get_or_404(po_id)
    return render_template('purchase/detail.html', po=po)


@purchase_bp.route('/<int:po_id>/submit', methods=['POST'])
@login_required
def submit(po_id):
    """提交审批"""
    success, msg = PurchaseService.submit_for_approval(po_id, current_user)
    if success:
        flash('已提交审批', 'success')
    else:
        flash(msg, 'danger')
    # 返回来源页面
    referer = request.referrer
    if referer and 'purchase' in referer and 'detail' not in referer:
        return redirect(referer)
    return redirect(url_for('purchase.detail', po_id=po_id))


@purchase_bp.route('/<int:po_id>/approve', methods=['POST'])
@login_required
def approve(po_id):
    """审批通过"""
    success, msg = PurchaseService.approve(po_id, current_user)
    if success:
        flash('审批通过', 'success')
    else:
        flash(msg, 'danger')
    # 返回来源页面
    referer = request.referrer
    if referer and 'purchase' in referer and 'detail' not in referer:
        return redirect(referer)
    return redirect(url_for('purchase.detail', po_id=po_id))


@purchase_bp.route('/<int:po_id>/reject', methods=['POST'])
@login_required
def reject(po_id):
    """审批拒绝"""
    reason = request.form.get('reason', '')
    # 使用 approve 方法，传入 approved=False 来拒绝
    success, msg = PurchaseService.approve(po_id, current_user, approved=False, remark=reason)
    if success:
        flash('已拒绝', 'success')
    else:
        flash(msg, 'danger')
    # 返回来源页面
    referer = request.referrer
    if referer and 'purchase' in referer and 'detail' not in referer:
        return redirect(referer)
    return redirect(url_for('purchase.detail', po_id=po_id))


@purchase_bp.route('/<int:po_id>/receive', methods=['GET', 'POST'])
@login_required
def receive(po_id):
    """收货入库"""
    po = PurchaseOrder.query.get_or_404(po_id)
    
    if po.status != PurchaseOrder.STATUS_APPROVED:
        flash('只有已审批的订单才能收货', 'danger')
        return redirect(url_for('purchase.detail', po_id=po_id))
    
    if request.method == 'POST':
        # 获取收货数据
        item_ids = request.form.getlist('item_id[]')
        received_qtys = request.form.getlist('received_qty[]')
        is_quality_passes = request.form.getlist('is_quality_pass[]')
        
        receives = []
        for i, item_id in enumerate(item_ids):
            if item_id and received_qtys[i]:
                receives.append({
                    'item_id': int(item_id),
                    'received_qty': int(received_qtys[i]),
                    'is_quality_pass': is_quality_passes[i] == '1' if i < len(is_quality_passes) else True
                })
        
        if not receives:
            flash('请输入收货数量', 'danger')
            return render_template('purchase/receive.html', po=po)
        
        success, msg = PurchaseService.receive_items(po_id, receives, current_user)
        if success:
            flash('收货成功', 'success')
            return redirect(url_for('purchase.detail', po_id=po_id))
        else:
            flash(msg, 'danger')
    
    return render_template('purchase/receive.html', po=po)


@purchase_bp.route('/<int:po_id>/cancel', methods=['POST'])
@login_required
def cancel(po_id):
    """取消订单"""
    reason = request.form.get('reason', '')
    po = PurchaseOrder.query.get_or_404(po_id)
    
    if po.status in [PurchaseOrder.STATUS_RECEIVED, PurchaseOrder.STATUS_CANCELLED]:
        flash('该订单无法取消', 'danger')
        referer = request.referrer
        if referer and 'purchase' in referer:
            return redirect(referer)
        return redirect(url_for('purchase.detail', po_id=po_id))
    
    po.status = PurchaseOrder.STATUS_CANCELLED
    if reason:
        po.remark = (po.remark or '') + f"\n[取消原因] {reason}"
    db.session.commit()
    
    flash('订单已取消', 'success')
    # 返回来源页面
    referer = request.referrer
    if referer and 'purchase' in referer and 'detail' not in referer:
        return redirect(referer)
    return redirect(url_for('purchase.detail', po_id=po_id))


@purchase_bp.route('/<int:po_id>/change-status', methods=['POST'])
@login_required
def change_status(po_id):
    """手动修改订单状态"""
    po = PurchaseOrder.query.get_or_404(po_id)
    new_status = request.form.get('new_status', '')
    reason = request.form.get('reason', '')
    
    # 验证状态值
    valid_statuses = ['draft', 'pending', 'approved', 'received', 'cancelled']
    if new_status not in valid_statuses:
        flash('无效的状态值', 'danger')
        return redirect(url_for('purchase.detail', po_id=po_id))
    
    # 如果状态相同则不做改变
    if po.status == new_status:
        flash('状态未变化', 'info')
        return redirect(url_for('purchase.detail', po_id=po_id))
    
    old_status = po.status
    po.status = new_status
    
    # 记录状态变更
    status_names = {
        'draft': '草稿',
        'pending': '待审批',
        'approved': '已审批',
        'received': '已收货',
        'cancelled': '已取消'
    }
    change_note = f"[状态变更] {status_names.get(old_status, old_status)} → {status_names.get(new_status, new_status)}"
    if reason:
        change_note += f" | 原因: {reason}"
    change_note += f" | 操作人: {current_user.username} | 时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    po.remark = (po.remark or '') + '\n' + change_note
    
    # 如果改为已审批状态，更新审批信息
    if new_status == 'approved':
        po.approved_at = datetime.utcnow()
        po.approved_by = current_user.id
    
    db.session.commit()
    
    flash(f'订单状态已更新为: {status_names.get(new_status, new_status)}', 'success')
    return redirect(url_for('purchase.detail', po_id=po_id))


# ============== API 接口 ==============

@purchase_bp.route('/api/product-price/<int:product_id>/<int:supplier_id>')
@login_required
def get_product_price(product_id, supplier_id):
    """获取商品采购价"""
    price = PurchaseService.get_supplier_price(product_id, supplier_id)
    return jsonify({'price': price})


@purchase_bp.route('/api/supplier-performance/<int:supplier_id>')
@login_required
def get_supplier_performance(supplier_id):
    """获取供应商绩效"""
    from app.models.purchase import SupplierPerformance
    
    perf = SupplierPerformance.query.filter_by(supplier_id=supplier_id).first()
    if not perf:
        return jsonify({'total_orders': 0, 'on_time_rate': 0, 'quality_rate': 0})
    
    return jsonify({
        'total_orders': perf.total_orders,
        'on_time_rate': round(perf.on_time_rate, 1),
        'quality_rate': round(perf.quality_rate, 1),
        'avg_delivery_days': perf.avg_delivery_days
    })

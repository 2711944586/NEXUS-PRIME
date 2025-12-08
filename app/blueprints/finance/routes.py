"""财务管理路由"""
from flask import render_template, request, flash, redirect, url_for, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from app.extensions import db
from app.blueprints.finance import finance_bp
from app.blueprints.finance.forms import CreditSettingForm, PaymentForm, StatementForm
from app.models.finance import CustomerCredit, Receivable, PaymentRecord, AccountStatement
from app.models.biz import Partner
from app.services.finance_service import FinanceService
from app.utils.decorators import permission_required


@finance_bp.route('/')
@login_required
def index():
    """财务管理首页"""
    # 应收账款统计
    aging = FinanceService.get_aging_analysis()
    total_receivable = sum(a['amount'] for a in aging.values())
    
    # 待收款订单数
    pending_count = Receivable.query.filter(
        Receivable.status.in_([Receivable.STATUS_PENDING, Receivable.STATUS_PARTIAL])
    ).count()
    
    # 逾期订单数
    overdue_count = Receivable.query.filter_by(status=Receivable.STATUS_OVERDUE).count()
    
    return render_template('finance/index.html',
                         aging=aging,
                         total_receivable=total_receivable,
                         pending_count=pending_count,
                         overdue_count=overdue_count)


# ============== 应收账款 ==============

@finance_bp.route('/receivables')
@login_required
def receivables():
    """应收账款列表"""
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    customer_id = request.args.get('customer_id', 0, type=int)
    search_query = request.args.get('q', '').strip()
    
    query = Receivable.query
    
    if status:
        query = query.filter_by(status=status)
    if customer_id:
        query = query.filter_by(customer_id=customer_id)
    
    # 搜索功能
    if search_query:
        query = query.join(Partner, Receivable.customer_id == Partner.id, isouter=True).filter(
            db.or_(
                Receivable.receivable_no.ilike(f'%{search_query}%'),
                Partner.name.ilike(f'%{search_query}%')
            )
        )
    
    pagination = query.order_by(Receivable.due_date.asc()).paginate(
        page=page, per_page=15, error_out=False
    )
    
    customers = Partner.query.filter(
        Partner.type.in_(['customer', 'both']),
        Partner.is_deleted == False
    ).all()
    
    # 统计数据
    from sqlalchemy import func
    total_receivable = db.session.query(func.sum(Receivable.total_amount)).scalar() or 0
    total_paid = db.session.query(func.sum(Receivable.paid_amount)).scalar() or 0
    pending_count = Receivable.query.filter(Receivable.status.in_(['pending', 'partial'])).count()
    overdue_count = Receivable.query.filter_by(status='overdue').count()
    
    return render_template('finance/receivables.html',
                         receivables=pagination.items,
                         pagination=pagination,
                         customers=customers,
                         current_status=status,
                         current_customer=customer_id,
                         total_receivable=total_receivable,
                         total_paid=total_paid,
                         pending_count=pending_count,
                         overdue_count=overdue_count)


@finance_bp.route('/receivables/<int:receivable_id>')
@login_required
def receivable_detail(receivable_id):
    """应收账款详情"""
    receivable = Receivable.query.get_or_404(receivable_id)
    payments = PaymentRecord.query.filter_by(receivable_id=receivable_id).order_by(PaymentRecord.payment_date.desc()).all()
    return render_template('finance/receivable_detail.html', receivable=receivable, payments=payments)


@finance_bp.route('/receivables/<int:receivable_id>/send_reminder', methods=['POST'])
@login_required
def send_reminder(receivable_id):
    """发送催款提醒"""
    from app.models.notification import Notification
    
    receivable = Receivable.query.get_or_404(receivable_id)
    
    # 检查状态
    if receivable.status not in ['pending', 'partial', 'overdue']:
        return jsonify({'success': False, 'message': '该应收款已结清，无需催款'})
    
    # 获取客户信息
    customer = receivable.customer
    if not customer:
        return jsonify({'success': False, 'message': '未找到客户信息'})
    
    # 创建催款通知记录
    notification = Notification(
        user_id=current_user.id,
        title=f'催款通知已发送 - {customer.name}',
        content=f'已向客户 {customer.name} 发送催款提醒。\n'
                f'应收单号: {receivable.receivable_no or "RCV-" + str(receivable.id)}\n'
                f'应收金额: ¥{receivable.total_amount:,.2f}\n'
                f'未收金额: ¥{receivable.unpaid_amount:,.2f}\n'
                f'到期日期: {receivable.due_date.strftime("%Y-%m-%d") if receivable.due_date else "-"}',
        type=Notification.TYPE_INFO,
        category='finance',
        related_type='receivable',
        related_id=receivable.id
    )
    db.session.add(notification)
    
    # 更新应收款的催款记录（如果有reminder_count字段的话）
    # 这里可以扩展模型添加催款次数、最后催款时间等字段
    
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': f'已向 {customer.name} 发送催款提醒',
        'customer_name': customer.name,
        'customer_phone': customer.phone or '-',
        'customer_email': customer.email or '-',
        'amount': f'¥{receivable.unpaid_amount:,.2f}'
    })


@finance_bp.route('/receivables/<int:receivable_id>/payment', methods=['GET', 'POST'])
@login_required
def record_payment(receivable_id):
    """记录收款"""
    receivable = Receivable.query.get_or_404(receivable_id)
    form = PaymentForm()
    form.receivable_id.data = receivable_id
    
    if form.validate_on_submit():
        success, result = FinanceService.record_payment(
            receivable_id=receivable_id,
            amount=float(form.amount.data),
            payment_method=form.payment_method.data,
            user=current_user,
            reference_no=form.reference_no.data,
            remark=form.remark.data
        )
        
        if success:
            flash(f'收款成功，单号: {result.payment_no}', 'success')
            return redirect(url_for('finance.receivable_detail', receivable_id=receivable_id))
        else:
            flash(result, 'danger')
    
    return render_template('finance/payment.html', form=form, receivable=receivable)


@finance_bp.route('/aging')
@login_required
def aging_report():
    """账龄分析报表"""
    customer_id = request.args.get('customer_id', 0, type=int)
    export_format = request.args.get('export', '')
    
    aging = FinanceService.get_aging_analysis(customer_id if customer_id else None)
    
    # 导出Excel
    if export_format == 'excel':
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "账龄分析报表"
        
        # 标题样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:D1')
        ws['A1'] = f"账龄分析报表 - {datetime.now().strftime('%Y年%m月%d日')}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表头
        headers = ['账龄区间', '笔数', '金额(元)', '占比(%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 数据
        rows_data = [
            ('未到期', aging['current']['count'], aging['current']['amount'], 0),
            ('逾期0-30天', aging['0-30']['count'], aging['0-30']['amount'], 0),
            ('逾期31-60天', aging['31-60']['count'], aging['31-60']['amount'], 0),
            ('逾期61-90天', aging['61-90']['count'], aging['61-90']['amount'], 0),
            ('逾期90天以上', aging['90+']['count'], aging['90+']['amount'], 0),
        ]
        
        total_amount = sum(r[2] for r in rows_data)
        
        for row_idx, (label, count, amount, _) in enumerate(rows_data, 4):
            pct = (amount / total_amount * 100) if total_amount > 0 else 0
            ws.cell(row=row_idx, column=1, value=label).border = thin_border
            ws.cell(row=row_idx, column=2, value=count).border = thin_border
            ws.cell(row=row_idx, column=3, value=f"¥{amount:,.2f}").border = thin_border
            ws.cell(row=row_idx, column=4, value=f"{pct:.1f}%").border = thin_border
        
        # 合计行
        total_row = len(rows_data) + 4
        ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=sum(r[1] for r in rows_data)).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=f"¥{total_amount:,.2f}").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="100%").font = Font(bold=True)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'账龄分析报表_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    customers = Partner.query.filter(
        Partner.type.in_(['customer', 'both']),
        Partner.is_deleted == False
    ).all()
    
    return render_template('finance/aging.html',
                         aging=aging,
                         customers=customers,
                         current_customer=customer_id,
                         now=datetime.now)


# ============== 客户信用 ==============

@finance_bp.route('/credits')
@login_required
def credits():
    """客户信用列表"""
    page = request.args.get('page', 1, type=int)
    filter_type = request.args.get('filter', '')
    per_page = 15
    
    # 基础查询
    query = CustomerCredit.query.join(Partner).filter(
        Partner.is_deleted == False
    )
    
    # 筛选条件
    if filter_type == 'frozen':
        # 冻结状态可以直接在数据库筛选
        query = query.filter(CustomerCredit.is_frozen == True)
        credits_list = query.order_by(CustomerCredit.used_credit.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif filter_type in ['normal', 'warning']:
        # 预警/正常需要在Python端筛选，因为is_warning是计算属性
        all_unfrozen = query.filter(CustomerCredit.is_frozen == False).order_by(
            CustomerCredit.used_credit.desc()
        ).all()
        
        if filter_type == 'warning':
            filtered_items = [c for c in all_unfrozen if c.is_warning]
        else:
            filtered_items = [c for c in all_unfrozen if not c.is_warning]
        
        # 手动分页
        total = len(filtered_items)
        start = (page - 1) * per_page
        end = start + per_page
        items = filtered_items[start:end]
        
        # 创建一个简单的分页对象
        class SimplePagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page if total > 0 else 1
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1
                self.next_num = page + 1
        
        credits_list = SimplePagination(items, page, per_page, total)
    else:
        # 全部
        credits_list = query.order_by(CustomerCredit.used_credit.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    # 统计数据
    from sqlalchemy import func
    total_limit = db.session.query(func.sum(CustomerCredit.credit_limit)).join(Partner).filter(
        Partner.is_deleted == False
    ).scalar() or 0
    
    # 获取所有信用记录来计算预警数量
    all_credits = CustomerCredit.query.join(Partner).filter(
        Partner.is_deleted == False, CustomerCredit.is_frozen == False
    ).all()
    warning_count = sum(1 for c in all_credits if c.is_warning)
    normal_count = sum(1 for c in all_credits if not c.is_warning)
    
    frozen_count = CustomerCredit.query.join(Partner).filter(
        Partner.is_deleted == False,
        CustomerCredit.is_frozen == True
    ).count()
    
    return render_template('finance/credits.html', 
                         credits=credits_list,
                         current_filter=filter_type,
                         total_limit=total_limit,
                         warning_count=warning_count,
                         normal_count=normal_count,
                         frozen_count=frozen_count)


@finance_bp.route('/credits/<int:customer_id>/freeze', methods=['GET', 'POST'])
@login_required
@permission_required('finance.credit')
def freeze_credit(customer_id):
    """冻结信用 - 独立页面"""
    customer = Partner.query.get_or_404(customer_id)
    credit = FinanceService.get_or_create_credit(customer_id)
    
    if request.method == 'POST':
        reason = request.form.get('reason', '')
        success, msg = FinanceService.freeze_credit(customer_id, reason, current_user)
        if success:
            flash('已冻结客户信用', 'success')
        else:
            flash(msg, 'danger')
        return redirect(url_for('finance.credits'))
    
    return render_template('finance/freeze_credit.html', customer=customer, credit=credit)


@finance_bp.route('/credits/<int:customer_id>/setting', methods=['GET', 'POST'])
@login_required
@permission_required('finance.credit')
def credit_setting(customer_id):
    """设置客户信用额度"""
    customer = Partner.query.get_or_404(customer_id)
    credit = FinanceService.get_or_create_credit(customer_id)
    
    form = CreditSettingForm()
    
    if form.validate_on_submit():
        credit.credit_limit = float(form.credit_limit.data)
        credit.warning_threshold = float(form.warning_threshold.data) if form.warning_threshold.data else 80
        db.session.commit()
        flash('信用额度已更新', 'success')
        return redirect(url_for('finance.credits'))
    
    # 预填充
    form.customer_id.data = customer_id
    form.credit_limit.data = credit.credit_limit
    form.warning_threshold.data = credit.warning_threshold
    
    return render_template('finance/credit_setting.html', form=form, customer=customer, credit=credit)


@finance_bp.route('/credits/<int:customer_id>/unfreeze', methods=['POST'])
@login_required
@permission_required('finance.credit')
def unfreeze_credit(customer_id):
    """解冻信用"""
    success, msg = FinanceService.unfreeze_credit(customer_id)
    if success:
        flash('已解冻客户信用', 'success')
    else:
        flash(msg, 'danger')
    return redirect(url_for('finance.credits'))


# ============== 对账单 ==============

@finance_bp.route('/statements')
@login_required
def statements():
    """对账单列表"""
    page = request.args.get('page', 1, type=int)
    customer_id = request.args.get('customer_id', 0, type=int)
    
    query = AccountStatement.query
    if customer_id:
        query = query.filter_by(customer_id=customer_id)
    
    pagination = query.order_by(AccountStatement.period_end.desc()).paginate(
        page=page, per_page=15, error_out=False
    )
    
    customers = Partner.query.filter(
        Partner.type.in_(['customer', 'both']),
        Partner.is_deleted == False
    ).all()
    
    return render_template('finance/statements.html',
                         statements=pagination.items,
                         pagination=pagination,
                         customers=customers,
                         current_customer=customer_id,
                         now=datetime.now)


@finance_bp.route('/statements/generate', methods=['GET', 'POST'])
@login_required
def generate_statement():
    """生成对账单"""
    form = StatementForm()
    
    customers = Partner.query.filter(
        Partner.type.in_(['customer', 'both']),
        Partner.is_deleted == False
    ).all()
    form.customer_id.choices = [(c.id, c.name) for c in customers]
    
    if form.validate_on_submit():
        success, result = FinanceService.generate_statement(
            customer_id=form.customer_id.data,
            period_start=form.period_start.data,
            period_end=form.period_end.data,
            user=current_user
        )
        
        if success:
            flash(f'对账单生成成功: {result.statement_no}', 'success')
            return redirect(url_for('finance.statement_detail', statement_id=result.id))
        else:
            flash(result, 'danger')
    
    # 默认日期
    today = datetime.now().date()
    form.period_end.data = today
    form.period_start.data = today.replace(day=1)
    
    return render_template('finance/generate_statement.html', form=form)


@finance_bp.route('/statements/<int:statement_id>')
@login_required
def statement_detail(statement_id):
    """对账单详情"""
    statement = AccountStatement.query.get_or_404(statement_id)
    
    # 获取期间内的订单和收款
    from app.models.trade import Order
    
    orders = Order.query.filter(
        Order.customer_id == statement.customer_id,
        Order.created_at >= statement.period_start,
        Order.created_at <= statement.period_end,
        Order.status.in_(['paid', 'shipped', 'done'])
    ).all()
    
    payments = PaymentRecord.query.filter(
        PaymentRecord.customer_id == statement.customer_id,
        PaymentRecord.payment_date >= statement.period_start,
        PaymentRecord.payment_date <= statement.period_end
    ).all()
    
    return render_template('finance/statement_detail.html',
                         statement=statement,
                         orders=orders,
                         payments=payments)


@finance_bp.route('/statements/<int:statement_id>/confirm', methods=['POST'])
@login_required
def confirm_statement(statement_id):
    """确认对账单"""
    statement = AccountStatement.query.get_or_404(statement_id)
    
    if statement.confirmed:
        flash('该对账单已经确认过了', 'warning')
    else:
        statement.confirmed = True
        statement.confirmed_at = datetime.utcnow()
        db.session.commit()
        flash('对账单已确认', 'success')
    
    return redirect(url_for('finance.statement_detail', statement_id=statement_id))


@finance_bp.route('/statements/<int:statement_id>/unconfirm', methods=['POST'])
@login_required
def unconfirm_statement(statement_id):
    """取消确认对账单"""
    statement = AccountStatement.query.get_or_404(statement_id)
    
    if not statement.confirmed:
        flash('该对账单尚未确认', 'warning')
    else:
        statement.confirmed = False
        statement.confirmed_at = None
        db.session.commit()
        flash('已取消确认', 'success')
    
    return redirect(url_for('finance.statement_detail', statement_id=statement_id))


# ============== API 接口 ==============

@finance_bp.route('/api/check-credit/<int:customer_id>/<float:amount>')
@login_required
def check_credit_api(customer_id, amount):
    """检查信用额度API"""
    is_allowed, msg = FinanceService.check_credit(customer_id, amount)
    return jsonify({'allowed': is_allowed, 'message': msg})


@finance_bp.route('/api/customer-credit/<int:customer_id>')
@login_required
def get_customer_credit(customer_id):
    """获取客户信用信息"""
    credit = FinanceService.get_or_create_credit(customer_id)
    return jsonify({
        'credit_limit': credit.credit_limit,
        'used_credit': credit.used_credit,
        'available_credit': credit.available_credit,
        'usage_rate': credit.usage_rate,
        'is_frozen': credit.is_frozen,
        'is_warning': credit.is_warning
    })

"""
数据导出服务
支持 Excel、CSV 等格式导出
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportService:
    """数据导出服务"""
    
    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]], 
        columns: List[Dict[str, str]],
        sheet_name: str = "Sheet1",
        title: str = "数据导出"
    ) -> BytesIO:
        """
        导出数据到 Excel
        
        Args:
            data: 数据列表 [{"field1": value1, "field2": value2}, ...]
            columns: 列定义 [{"field": "field1", "header": "字段1", "width": 15}, ...]
            sheet_name: 工作表名称
            title: 报表标题
            
        Returns:
            BytesIO: Excel 文件流
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 样式定义
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
        cell_font = Font(name='微软雅黑', size=10)
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # 写入标题（合并单元格）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 写入导出时间
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        time_cell = ws.cell(row=2, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        time_cell.font = Font(name='微软雅黑', size=9, color='6B7280')
        time_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20
        
        # 写入表头
        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_def['header'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # 设置列宽
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = col_def.get('width', 15)
        
        ws.row_dimensions[3].height = 25
        
        # 写入数据
        for row_idx, row_data in enumerate(data, start=4):
            for col_idx, col_def in enumerate(columns, start=1):
                field = col_def['field']
                value = row_data.get(field, '')
                
                # 处理特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.border = border
                
                # 数字右对齐，其他左对齐
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.row_dimensions[row_idx].height = 20
        
        # 冻结前三行（标题 + 时间 + 表头）
        ws.freeze_panes = 'A4'
        
        # 保存到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]], 
        columns: List[Dict[str, str]]
    ) -> BytesIO:
        """
        导出数据到 CSV
        
        Args:
            data: 数据列表
            columns: 列定义
            
        Returns:
            BytesIO: CSV 文件流
        """
        import io
        
        # 使用 StringIO 写入 CSV
        text_output = io.StringIO()
        
        writer = csv.DictWriter(
            text_output, 
            fieldnames=[col['field'] for col in columns],
            extrasaction='ignore'
        )
        
        # 写入表头
        header_row = {col['field']: col['header'] for col in columns}
        writer.writerow(header_row)
        
        # 写入数据
        for row in data:
            processed_row = {}
            for col in columns:
                field = col['field']
                value = row.get(field, '')
                
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                
                processed_row[field] = value
            
            writer.writerow(processed_row)
        
        # 转换为 BytesIO（UTF-8 with BOM）
        output = BytesIO()
        output.write('\ufeff'.encode('utf-8'))
        output.write(text_output.getvalue().encode('utf-8'))
        output.seek(0)
        
        return output


# 全局单例
export_service = ExportService()
